import openpyxl
import uvicorn
import json
import sqlite3
from config import *
from fastapi import FastAPI

dataframe = openpyxl.load_workbook("llll.xlsx")
dataframe1 = dataframe["01.04-06.04"]
sqlite_conn = sqlite3.connect("base.db", check_same_thread=False)
cursor = sqlite_conn.cursor()

#TODO remap class map and make no lesson diaspear

app = FastAPI()

breaks = ["Перемена 10 минут", "Перемена 30 минут", "Перемена 30 минут", "Перемена 10 минут", "Перемена 10 минут",
          "Конец пар"]


@app.get("/")
def read_root():
    return {"Hello": "World"}


def get_groups():
    groups = []
    # getting all cell values in whole line 6
    for cell in dataframe1[y_offset]:
        if cell.value is not None and cell.value not in filter_words:
            groups.append(cell.value)
    return groups


@app.get("/groups")
def return_groups():
    groups = get_groups()
    return {"groups": groups}


@app.get("/insert_token/{couple}")
def insert_token(couple: str):
    fcm_token, group = couple.split("|")

    in_db = cursor.execute("SELECT COUNT(*) FROM users WHERE fcm_token = ?", (fcm_token,)).fetchone()

    if in_db[0] == 0:
        cursor.execute("INSERT INTO users VALUES (?, ?)", (fcm_token, group))
        sqlite_conn.commit()
    else:
        cursor.execute("UPDATE users SET group_name = ? WHERE fcm_token = ?", (group, fcm_token))
        sqlite_conn.commit()
    return {"status": "ok"}


@app.get("/timetable/{group}")
def get_lessons(group: str):
    classes = {}
    counter = 0
    day_counter = -1
    all_lessons = []

    for row in dataframe1[group_map[group] + "6": group_map[group] + "43"]:
        for cell in row:
            class_to_write = str(cell.value).replace("None", "Нет пары")
            replace_class = dataframe1.cell(row=cell.row, column=cell.column + 1).value
            class_room_number = str(dataframe1.cell(row=cell.row, column=cell.column + 2).value).replace(".0",
                                                                                                         "")

            if replace_class is not None:
                class_to_write += " | " + str(replace_class)

            all_lessons.append(class_to_write + " ~ " + class_room_number)

    for lesson in all_lessons[2:]:
        if counter == 0 or counter == 6:
            counter = 0
            day_counter += 1
            classes[day_counter] = []

        classes[day_counter].append(lesson)
        # print(lesson)

        counter += 1

    for lesson in classes:
        classes[lesson] = [val for pair in zip(classes[lesson], breaks) for val in pair]

    for day in classes:
        if "Нет пары" in str(classes[day]):
            for elem in classes[day]:
                if "Нет пары" in elem:
                    idx = classes[day].index(elem)
                    break
            classes[day] = classes[day][:idx]
        else:
            pass

    for day in classes:
        if "Перемена" in classes[day][-1]:
            classes[day] = classes[day][:-1]

    return classes


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
