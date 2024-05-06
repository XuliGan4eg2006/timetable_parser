import openpyxl
from config import *

dataframe = openpyxl.load_workbook("llll.xlsx")
dataframe1 = dataframe["01.04-06.04"]

#TODO remap config group map

breaks = ["Перемена 10 минут", "Перемена 30 минут", "Перемена 30 минут", "Перемена 10 минут", "Перемена 10 минут",
          "Конец пар"]


def get_groups():
    groups = []
    for cell in dataframe1[y_offset]:
        if cell.value is not None and cell.value not in filter_words:
            groups.append(cell.value)
    return groups


def get_lessons(group: str):
    classes = {}
    counter = 0
    day_counter = 0
    all_lessons = []

    for row in dataframe1[group_map[group] + "6": group_map[group] + "43"]:
        for cell in row:

            class_to_write = str(cell.value).replace("None", "Нет урока")

            replace_class = dataframe1.cell(row=cell.row, column=cell.column + 1).value
            class_room_number = str(dataframe1.cell(row=cell.row, column=cell.column + 2).value).replace(".0",
                                                                                                         "")

            if replace_class is not None:
                class_to_write += " | " + str(replace_class)

            all_lessons.append(class_to_write + " ~ " + class_room_number)

    for lesson in all_lessons[2:]:
        if counter == 0 or counter == 6:
            counter = 0
            day_counter += 1
            classes[day_counter] = []

        classes[day_counter].append(lesson)
        print(lesson)

        counter += 1

    for lesson in classes:
        classes[lesson] = [val for pair in zip(classes[lesson], breaks) for val in pair]

    return classes


print(get_lessons("ИСП9-123А"))
