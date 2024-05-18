import openpyxl
import redis
import requests
from tqdm import tqdm
from config import *


def get_lessons(dataframe, group):
    classes = {}
    counter = 0
    day_counter = -1
    all_lessons = []

    for row in dataframe1[group_map[group] + "6": group_map[group] + "43"]:
        for cell in row:
            class_to_write = str(cell.value).replace("None", "Нет пары")
            replace_class = dataframe1.cell(row=cell.row, column=cell.column + 1).value
            class_room_number = str(dataframe1.cell(row=cell.row, column=cell.column + 2).value).replace(".0",
                                                                                                         "")

            if replace_class is not None:
                class_to_write += " | " + str(replace_class)

            all_lessons.append(class_to_write + " ~ " + class_room_number)

    for lesson in all_lessons[2:]:
        if counter == 0 or counter == 6:
            counter = 0
            day_counter += 1
            classes[day_counter] = []

        classes[day_counter].append(lesson)

        counter += 1

    for lesson in classes:
        classes[lesson] = [val for pair in zip(classes[lesson], breaks) for val in pair]

    for day in classes:
        if "Нет пары" in str(classes[day]):
            for elem in classes[day]:
                if "Нет пары" in elem:
                    idx = classes[day].index(elem)
                    break
            classes[day] = classes[day][:idx]
        else:
            pass

    for day in classes:
        if "Перемена" in classes[day][-1]:
            classes[day] = classes[day][:-1]
    return classes


def init_db():
    dataframe = openpyxl.load_workbook("llll.xlsx")
    dataframe1 = dataframe["01.04-06.04"]
    r = redis.Redis(host='localhost', port=6379, db=0)
    print("Initing db.....")
    for group in tqdm(group_map):
        classes = str(get_lessons(dataframe1, group))
        r.set(group, classes)
    print("Done")


def send_notification():
    data = {
        "message": {
            "token": "fjrd0ZzRRcyC4EhgYDdnlB:APA91bHt1NBi6kIGpE3GGIVqCWgQFmm8aPczmjReT-2kPkGThSBi_zDFZWyQv9eXYbYoOWVuOhakUuKi3M-jb5X5MU_ZLgaz0w038VWAzGhWKxixp6sbl62uYPgshCnl1V2B_5oepJIZ",
            "notification": {
                "body": "This is an FCM notification message!",
                "title": "FCM Message"
            }
        }
    }
    headers = {
        "Content-Type": "application/json",
        "Authorization": "Bearer ya29.a0AXooCgvhPMNLUsH1CN0iyzWqaKxMhiqzE0fIiDqj6-J7JBE4HCQdPJg5YwvV0h4j-xaeexpZ9pEMefhBRIXlZZRUpx5HNJJ35MLo_mCznsKd10CmrAEcMg1CNJjwRtPFUOxkPa_y-hQs7sT8bZpEgfwpwvcikBllIe0qaCgYKAZwSARESFQHGX2MiHLVftXAmueREbMjJ25yWFA0171"
    }
    req = requests.post("https://fcm.googleapis.com/v1/projects/mtusi-timetable/messages:send", json=data, headers=headers)

    print(req.text)

send_notification()
