import ast
import os
import sqlite3
import json
import openpyxl
import redis
import psycopg2
import time

import requests

from fcm import send_notification

r = redis.Redis(host='redis', port=6379, db=0)
print("Initing db.....")

psycopg_connection = psycopg2.connect(dbname='mtusi', user='postgres', password='lmaooo123',
                                      host='db',
                                      port='5432')
cursor = psycopg_connection.cursor()

selected_page = r.get("selected_page").decode('utf-8')
download_url = r.get("download_url").decode('utf-8')
group_map = json.loads('"' + r.get("group_map").decode('utf-8') + '"')
breaks = r.get("breaks").decode('utf-8').split(",")

def download_sheet():
    req = requests.get(download_url)
    if req.status_code == 200:
        with open("temp.xlsx", "wb") as f:
            f.write(req.content)
        return True
    else:
        print("Error downloading sheet")
        return False


def get_lessons(dataframe, group):
    classes = {}
    counter = 0
    day_counter = -1
    all_lessons = []

    for row in dataframe[group_map[group] + "6": group_map[group] + "43"]:
        for cell in row:
            class_to_write = str(cell.value).replace("None", "Нет пары")
            replace_class = dataframe.cell(row=cell.row, column=cell.column + 1).value
            class_room_number = str(dataframe.cell(row=cell.row, column=cell.column + 2).value).replace(".0",
                                                                                                        "")

            if replace_class is not None:
                class_to_write += " | " + str(replace_class)

            all_lessons.append(class_to_write + " ~ " + class_room_number)

    for lesson in all_lessons[2:]: #filling lessons
        if counter == 0 or counter == 6:
            counter = 0
            day_counter += 1
            classes[day_counter] = []

        classes[day_counter].append(lesson)

        counter += 1

    for lesson in classes: #filling breaks
        classes[lesson] = [val for pair in zip(classes[lesson], breaks) for val in pair]

    for day in classes: #removing no lessons in end
        if "Нет пары" in str(classes[day]):
            for elem in classes[day]:
                if "Нет пары" in elem:
                    idx = classes[day].index(elem)
                    break
            classes[day] = classes[day][:idx]
        else:
            pass

    # if not classes[day_counter]:
    #     classes[day_counter].append("No lessons")

    # for day in classes:
    #     if "Перемена" in classes[day][-1]:
    #         classes[day] = classes[day][:-1]

    return classes


while True:
    # download_sheet()
    # print("Done")
    #
    # dataframe = openpyxl.load_workbook("temp.xlsx")
    # dataframe1 = dataframe[selected_page]
    #
    # for group in group_map:
    #     old = r.get(group)
    #     old_dict = ast.literal_eval(old.decode('utf-8'))
    #
    #     new = get_lessons(dataframe1, group)
    #     if old_dict != new:
    #         cursor.execute("SELECT fcm_token FROM users WHERE group_name = %s", (group,))
    #         tokens = cursor.fetchall()
    #         if len(tokens) > 0:
    #             for token in tokens:
    #                 print("Sending notification to " + token[0])
    #                 send_notification(token[0], "Есть изменения в расписании",
    #                                   f"Для группы {group} есть изменения в расписании")
    #             r.set(group, str(new))
    #             print("Updated " + group)
    # os.remove("temp.xlsx")
    print("Sleeping...")
    time.sleep(1800)
