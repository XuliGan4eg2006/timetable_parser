import openpyxl
import redis
import requests
import os
from tqdm import tqdm
from config import *

r = redis.Redis(host='redis', port=6379, db=0)


def get_lessons(dataframe, group):
    classes = {}
    counter = 0
    day_counter = -1
    all_lessons = []

    for row in dataframe[group_map[group] + "6": group_map[group] + "43"]:
        for cell in row:
            class_to_write = str(cell.value).replace("None", "Нет пары")
            replace_class = dataframe.cell(row=cell.row, column=cell.column + 1).value
            class_room_number = str(dataframe.cell(row=cell.row, column=cell.column + 2).value).replace(".0",
                                                                                                        "")

            if replace_class is not None:
                class_to_write += " | " + str(replace_class)

            all_lessons.append(class_to_write + " ~ " + class_room_number)

    for lesson in all_lessons[2:]:
        if counter == 0 or counter == 6:
            counter = 0
            day_counter += 1
            classes[day_counter] = []

        classes[day_counter].append(lesson)

        counter += 1

    for lesson in classes:
        classes[lesson] = [val for pair in zip(classes[lesson], breaks) for val in pair]

    for day in classes:
        if "Нет пары" in str(classes[day]):
            for elem in classes[day]:
                if "Нет пары" in elem:
                    idx = classes[day].index(elem)
                    break
            classes[day] = classes[day][:idx]
        else:
            pass

    # for day in classes:
    #     if "Перемена" in classes[day][-1]:
    #         classes[day] = classes[day][:-1]

    return classes


def update_db():
    dataframe = openpyxl.load_workbook("xlsxs/llll.xlsx")
    dataframe1 = dataframe[selected_page]

    for group in tqdm(group_map):
        classes = str(get_lessons(dataframe1, group))
        r.set(group, classes)

    print("Done")


def download_sheet(new_sheet_name="llll.xlsx"):
    req = requests.get(download_url)
    if req.status_code == 200:

        if new_sheet_name == "llll.xlsx":
            try:
                os.remove("xlsxs/llll.xlsx")
            except:
                pass

        with open("xlsxs/" + new_sheet_name, "wb") as f:
            f.write(req.content)
        return True
    else:
        print("Error downloading sheet")
        return False


def insert_groups():
    r.set("groups", str(groups).replace("[", "").replace("]", ""))
